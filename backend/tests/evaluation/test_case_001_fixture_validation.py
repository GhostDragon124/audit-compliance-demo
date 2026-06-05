"""
Phase 0: Case 001 Fixture Validation

Validates that the Case 001 evaluation fixture directory is complete and correct.
Must pass WITHOUT real Qwen, Embedding, ChromaDB, OCR, or network.
"""

import json
import yaml
from pathlib import Path
import pytest
import PIL.Image

# ── Helpers ──────────────────────────────────────────────────────────

CASE_ROOT = (
    Path(__file__).resolve().parents[3]
    / "data"
    / "tests"
    / "evaluation"
    / "case_001_procurement_pending_approval"
)

REQUIRED_FILES = [
    "README.md",
    "manifest.json",
    "question.txt",
    "expected_results.yaml",
]


def test_case_root_exists():
    """1. Case root directory exists."""
    assert CASE_ROOT.is_dir(), (
        f"Case root not found at {CASE_ROOT}"
    )


def test_required_root_files_exist():
    """2. All required root files exist."""
    for fname in REQUIRED_FILES:
        path = CASE_ROOT / fname
        assert path.is_file(), f"Missing required file: {fname}"


# ── Manifest integrity ──────────────────────────────────────────────


def load_manifest():
    path = CASE_ROOT / "manifest.json"
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def test_manifest_is_valid_json():
    """manifest.json parses as valid JSON."""
    manifest = load_manifest()
    assert isinstance(manifest, dict)


def test_manifest_classification():
    """manifest classification is synthetic_non_sensitive."""
    manifest = load_manifest()
    assert manifest.get("classification") == "synthetic_non_sensitive"


def test_manifest_materials_exist():
    """All materials referenced in manifest exist on disk."""
    manifest = load_manifest()
    for rel_path in manifest.get("materials", []):
        resolved = (CASE_ROOT / rel_path).resolve()
        assert str(resolved).startswith(str(CASE_ROOT.resolve())), (
            f"Path traversal: {rel_path}"
        )
        assert resolved.is_file(), f"Material not found: {rel_path}"


def test_manifest_regulations_exist():
    """All regulations referenced in manifest exist on disk."""
    manifest = load_manifest()
    for rel_path in manifest.get("regulations", []):
        resolved = (CASE_ROOT / rel_path).resolve()
        assert str(resolved).startswith(str(CASE_ROOT.resolve())), (
            f"Path traversal: {rel_path}"
        )
        assert resolved.is_file(), f"Regulation not found: {rel_path}"


def test_manifest_ground_truth_exist():
    """All ground_truth files referenced in manifest exist on disk."""
    manifest = load_manifest()
    for rel_path in manifest.get("ground_truth", []):
        resolved = (CASE_ROOT / rel_path).resolve()
        assert str(resolved).startswith(str(CASE_ROOT.resolve())), (
            f"Path traversal: {rel_path}"
        )
        assert resolved.is_file(), f"Ground truth not found: {rel_path}"


def test_manifest_no_path_traversal():
    """Manifest paths do not escape the case directory."""
    manifest = load_manifest()
    for key in ("materials", "regulations", "ground_truth"):
        for rel_path in manifest.get(key, []):
            resolved = (CASE_ROOT / rel_path).resolve()
            assert str(resolved).startswith(str(CASE_ROOT.resolve())), (
                f"Path traversal detected in {key}: {rel_path}"
            )


# ── Expected results structure ──────────────────────────────────────


def load_expected():
    path = CASE_ROOT / "expected_results.yaml"
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def test_question_txt_nonempty():
    """question.txt exists and is non-empty."""
    question = (CASE_ROOT / "question.txt").read_text(encoding="utf-8").strip()
    assert len(question) > 0, "question.txt is empty"


def test_expected_results_yaml_valid():
    """expected_results.yaml is valid YAML."""
    data = load_expected()
    assert isinstance(data, dict)


def test_expected_results_has_required_keys():
    """expected_results.yaml has all required top-level keys."""
    data = load_expected()
    required_keys = [
        "case_id",
        "question",
        "materials",
        "regulations",
        "expected_ocr_fields",
        "retrieval_expectations",
        "expected_risk_points",
        "expected_uncertainties",
        "must_include_concepts",
        "must_not_claim",
        "pass_criteria",
    ]
    for key in required_keys:
        assert key in data, f"Missing expected key: {key}"


def test_expected_results_regulations_no_overlap():
    """relevant and irrelevant_distractor regulation lists do not overlap."""
    data = load_expected()
    relevant = set(data.get("regulations", {}).get("relevant", []))
    distractor = set(data.get("regulations", {}).get("irrelevant_distractor", []))
    overlap = relevant & distractor
    assert len(overlap) == 0, (
        f"Overlap between relevant and irrelevant_distractor: {overlap}"
    )


def test_expected_risk_points_nonempty():
    """expected_risk_points is non-empty."""
    data = load_expected()
    risks = data.get("expected_risk_points", [])
    assert len(risks) > 0, "expected_risk_points is empty"


def test_expected_ocr_fields_nonempty():
    """expected_ocr_fields is non-empty."""
    data = load_expected()
    fields = data.get("expected_ocr_fields", [])
    assert len(fields) > 0, "expected_ocr_fields is empty"


def test_must_not_claim_nonempty():
    """must_not_claim is non-empty."""
    data = load_expected()
    claims = data.get("must_not_claim", [])
    assert len(claims) > 0, "must_not_claim is empty"


# ── Text materials readability ──────────────────────────────────────


def test_material_purchase_declaration_content():
    """采购项目说明.md contains expected keywords."""
    path = CASE_ROOT / "materials" / "采购项目说明.md"
    content = path.read_text(encoding="utf-8")
    assert "实验室服务器采购项目" in content
    assert "850,000" in content or "850000" in content
    assert "询价采购" in content
    assert "星海科技有限公司" in content
    assert "审批中" in content or "审批流程" in content


def test_material_contract_draft_content():
    """合同草案.txt contains expected keywords."""
    path = CASE_ROOT / "materials" / "合同草案.txt"
    content = path.read_text(encoding="utf-8")
    assert "实验室服务器采购合同" in content
    assert "星海科技有限公司" in content
    assert "850,000" in content or "850000" in content


def test_material_supplier_quote_xlsx_readable():
    """供应商报价.xlsx is readable via openpyxl."""
    import openpyxl

    path = CASE_ROOT / "materials" / "供应商报价.xlsx"
    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) >= 1
    ws = wb.active
    assert ws is not None, "XLSX has no active sheet"
    assert ws.max_row is not None and ws.max_row >= 1


def test_regulation_files_readable():
    """All three regulation TXT files are readable."""
    for fname in ["采购管理办法.txt", "采购审批管理规定.txt", "差旅费管理规定.txt"]:
        path = CASE_ROOT / "regulations" / fname
        content = path.read_text(encoding="utf-8")
        assert len(content.strip()) > 0, f"{fname} is empty"


# ── Image validity ──────────────────────────────────────────────────


def test_image_is_valid():
    """采购审批单_scan.png is a valid image (non-zero size, correct format)."""
    path = CASE_ROOT / "materials" / "采购审批单_scan.png"
    assert path.is_file(), "PNG file not found"
    assert path.stat().st_size > 0, "PNG file is zero bytes"
    img = PIL.Image.open(path)
    img.verify()  # raises on corruption
    assert img.format == "PNG", f"Expected PNG format, got {img.format}"
    img.close()


# ── Ground truth ────────────────────────────────────────────────────


def test_ground_truth_exists():
    """采购审批单_ground_truth.txt exists and is readable."""
    path = CASE_ROOT / "ground_truth" / "采购审批单_ground_truth.txt"
    assert path.is_file()
    content = path.read_text(encoding="utf-8")
    assert len(content.strip()) > 0


def test_ground_truth_contains_expected_fields():
    """Ground truth contains expected key fields."""
    path = CASE_ROOT / "ground_truth" / "采购审批单_ground_truth.txt"
    content = path.read_text(encoding="utf-8")
    for field in [
        "实验室服务器采购项目",
        "850,000",
        "询价采购",
        "待审批",
        "未签字",
        "审批中",
    ]:
        assert field in content, f"Ground truth missing field: {field}"


# ── Case data isolation ─────────────────────────────────────────────


def test_case_regulations_not_in_production_regulations():
    """Case regulations are NOT in data/regulations/raw/."""
    prod_reg_dir = CASE_ROOT.parents[2] / "regulations" / "raw"
    if not prod_reg_dir.is_dir():
        pytest.skip("Production regulations directory does not exist")
    case_reg_names = {
        "采购管理办法.txt",
        "采购审批管理规定.txt",
        "差旅费管理规定.txt",
    }
    for f in prod_reg_dir.iterdir():
        if f.is_file():
            assert f.name not in case_reg_names, (
                f"Case regulation {f.name} found in production regulations"
            )
